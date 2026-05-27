"""Extract full BPPU REF data."""
import openpyxl

base = "d:\\Dani's Projects\\AUTOMATION (By Dani)\\Coretax Helper\\Sample Data\\Data Format and Template From Coretax"
path = f"{base}\\BPPU Excel to XML v.3.xlsx"
wb = openpyxl.load_workbook(path)

ws = wb['REF']
print(f"REF: {ws.max_row} rows")

# Tax objects (col A-C)
print("\n=== BPPU Tax Objects ===")
for r in range(2, ws.max_row + 1):
    code = ws.cell(row=r, column=1).value
    name = ws.cell(row=r, column=2).value
    rate = ws.cell(row=r, column=3).value
    if code:
        print(f'    ("{code}", "{name}", {rate}),')

# Facilities (col F-G)  
print("\n=== BPPU Facilities ===")
for r in range(2, ws.max_row + 1):
    code = ws.cell(row=r, column=6).value
    name = ws.cell(row=r, column=7).value
    if code and not code.startswith("Kode"):
        print(f'    "{code}": "{name}",')

# Documents (col F-G after row 18ish)
# Already in the output above
