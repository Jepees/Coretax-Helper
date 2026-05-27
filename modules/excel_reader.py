"""
Excel reader for Coretax BP21, BPMP, BPA1, and BPPU templates.
Reads the DATA sheet from uploaded Excel files and extracts structured data.
"""

import openpyxl
from datetime import datetime, date
from io import BytesIO

from modules.tax_engine import calculate_rate, get_deemed, get_tariff_type, get_bppu_rate
from modules.reference_data import (
    BP21_TAX_OBJECT_LOOKUP,
    BPMP_TAX_OBJECT_LOOKUP,
    BPA1_TAX_OBJECT_LOOKUP,
    BPPU_TAX_OBJECT_LOOKUP,
)


def _safe_str(value) -> str:
    """Convert value to string safely, handling None."""
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _safe_float(value) -> float:
    """Convert value to float safely."""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def _safe_date(value) -> str:
    """Convert value to YYYY-MM-DD string."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _is_formula(value) -> bool:
    """Check if a cell value is a formula string."""
    return isinstance(value, str) and value.startswith("=")


def read_bp21(file_bytes: bytes) -> dict:
    """
    Read BP21 Excel template.

    Expected structure:
    - Sheet "DATA" with Table1 (range B3:P9+)
    - Row 1: A1:B1 merged = label "NPWP Pemotong", C1 = NPWP value
    - Row 3: column headers (B=Masa Pajak, C=Tahun Pajak, ...)
    - Row 4+: data rows (columns B-P)

    Returns:
        dict with:
            - "tin": NPWP Pemotong (str)
            - "rows": list of dicts with XML field names as keys
            - "errors": list of validation error strings
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=False)

    if "DATA" not in wb.sheetnames:
        return {
            "tin": "",
            "rows": [],
            "errors": ["Sheet 'DATA' tidak ditemukan dalam file Excel."],
        }

    ws = wb["DATA"]
    errors = []
    rows = []

    # Find data start row — look for the table header row
    # BP21: Row 3 is typically the header, data starts at row 4
    data_start_row = 4
    header_row = 3

    # Verify headers exist
    header_b = ws.cell(row=header_row, column=2).value
    if header_b is None:
        # Try to find the header row by searching
        for r in range(1, 10):
            if ws.cell(row=r, column=2).value is not None:
                header_row = r
                data_start_row = r + 1
                break

    # Extract TIN from cell C1 (label "NPWP Pemotong" is in merged A1:B1)
    tin = _safe_str(ws.cell(row=1, column=3).value)
    if not tin:
        # Fallback: search row 1-2 for a 16-digit number
        for r in range(1, 4):
            for c in range(1, 6):
                val = _safe_str(ws.cell(row=r, column=c).value)
                if len(val) == 16 and val.isdigit():
                    tin = val
                    break
            if tin:
                break

    if not tin:
        errors.append("NPWP Pemotong tidak ditemukan (diharapkan di cell C1).")

    # Read data rows
    # Column mapping for BP21 DATA sheet:
    # A = NPWP Pemotong (TIN) — same for all rows
    # B = Masa Pajak (TaxPeriodMonth)
    # C = Tahun Pajak (TaxPeriodYear)
    # D = NPWP (CounterpartTin)
    # E = ID TKU Penerima (IDPlaceOfBusinessActivityOfIncomeRecipient)
    # F = Status PTKP (StatusTaxExemption)
    # G = Fasilitas (TaxCertificate)
    # H = Kode Objek Pajak (TaxObjectCode)
    # I = Penghasilan (Gross)
    # J = Deemed (formula VLOOKUP — needs recalc)
    # K = Tarif (formula — needs recalc)
    # L = Jenis Dok. Referensi (Document)
    # M = Nomor Dok. Referensi (DocumentNumber)
    # N = Tanggal Dok. Referensi (DocumentDate)
    # O = ID TKU Pemotong (IDPlaceOfBusinessActivity)
    # P = Tanggal Pemotongan (WithholdingDate)

    for r in range(data_start_row, ws.max_row + 1):
        # Check if row has data (at least Masa Pajak or NPWP filled)
        masa = ws.cell(row=r, column=2).value
        npwp_penerima = ws.cell(row=r, column=4).value
        if masa is None and npwp_penerima is None:
            continue

        tax_object_code = _safe_str(ws.cell(row=r, column=8).value)
        ptkp_status = _safe_str(ws.cell(row=r, column=6).value)
        gross = _safe_float(ws.cell(row=r, column=9).value)

        # Get deemed — recalculate from tax object code (don't rely on formula)
        deemed_cell = ws.cell(row=r, column=10).value
        if _is_formula(deemed_cell) or deemed_cell is None:
            deemed = get_deemed(tax_object_code)
        else:
            deemed = _safe_float(deemed_cell)

        # Calculate rate — always recalculate (formula cells won't have values)
        rate_cell = ws.cell(row=r, column=11).value
        if _is_formula(rate_cell) or rate_cell is None:
            rate = calculate_rate(tax_object_code, ptkp_status, gross, deemed, "BP21")
        else:
            rate = _safe_float(rate_cell)

        row_data = {
            "TaxPeriodMonth": int(_safe_float(masa)) if masa else 0,
            "TaxPeriodYear": int(_safe_float(ws.cell(row=r, column=3).value)) if ws.cell(row=r, column=3).value else 0,
            "CounterpartTin": _safe_str(npwp_penerima),
            "IDPlaceOfBusinessActivityOfIncomeRecipient": _safe_str(ws.cell(row=r, column=5).value),
            "StatusTaxExemption": ptkp_status,
            "TaxCertificate": _safe_str(ws.cell(row=r, column=7).value),
            "TaxObjectCode": tax_object_code,
            "Gross": gross,
            "Deemed": deemed,
            "Rate": rate,
            "Document": _safe_str(ws.cell(row=r, column=12).value),
            "DocumentNumber": _safe_str(ws.cell(row=r, column=13).value),
            "DocumentDate": _safe_date(ws.cell(row=r, column=14).value),
            "IDPlaceOfBusinessActivity": _safe_str(ws.cell(row=r, column=15).value),
            "WithholdingDate": _safe_date(ws.cell(row=r, column=16).value),
        }

        # Validate
        row_num = r
        if len(row_data["CounterpartTin"]) != 16 and row_data["CounterpartTin"]:
            errors.append(f"Baris {row_num}: NPWP/NIK penerima harus 16 digit (ditemukan: {len(row_data['CounterpartTin'])} digit)")
        if tax_object_code and tax_object_code not in BP21_TAX_OBJECT_LOOKUP:
            errors.append(f"Baris {row_num}: Kode objek pajak '{tax_object_code}' tidak dikenali")

        rows.append(row_data)

    if not rows:
        errors.append("Tidak ada data yang ditemukan di sheet DATA.")

    return {"tin": tin, "rows": rows, "errors": errors}


def read_bpmp(file_bytes: bytes) -> dict:
    """
    Read BPMP Excel template.

    Expected structure:
    - Sheet "DATA" with Table2
    - Cell A3: NPWP Pemotong
    - Row 4: column headers for the table (B-N)
    - Row 5+: data rows
    - Column B = Masa Pajak, C = Tahun Pajak, D = Status Pegawai, etc.

    Returns:
        dict with:
            - "tin": NPWP Pemotong (str)
            - "rows": list of dicts with XML field names as keys
            - "errors": list of validation error strings
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=False)

    if "DATA" not in wb.sheetnames:
        return {
            "tin": "",
            "rows": [],
            "errors": ["Sheet 'DATA' tidak ditemukan dalam file Excel."],
        }

    ws = wb["DATA"]
    errors = []
    rows = []

    # BPMP: TIN is in cell B1 (label "NPWP Pemotong" in A1, value in B1)
    tin = _safe_str(ws.cell(row=1, column=2).value)
    if not tin:
        # Fallback: search first few rows for a 16-digit number
        for r in range(1, 6):
            for c in range(1, 6):
                val = _safe_str(ws.cell(row=r, column=c).value)
                if len(val) == 16 and val.isdigit():
                    tin = val
                    break
            if tin:
                break

    if not tin:
        errors.append("NPWP Pemotong tidak ditemukan (diharapkan di cell B1).")

    # Data starts at row 5 (row 4 = headers)
    data_start_row = 5

    # Column mapping for BPMP DATA sheet:
    # B (2) = Masa Pajak (TaxPeriodMonth)
    # C (3) = Tahun Pajak (TaxPeriodYear)
    # D (4) = Status Pegawai (CounterpartOpt)
    # E (5) = NPWP/NIK/TIN (CounterpartTin)
    # F (6) = Nomor Passport (CounterpartPassport)
    # G (7) = Status PTKP (StatusTaxExemption)
    # H (8) = Posisi (Position)
    # I (9) = Sertifikat/Fasilitas (TaxCertificate)
    # J (10) = Kode Objek Pajak (TaxObjectCode)
    # K (11) = Penghasilan Kotor (Gross)
    # L (12) = Tarif (Rate) — formula, needs recalc
    # M (13) = ID TKU (IDPlaceOfBusinessActivity)
    # N (14) = Tanggal Pemotongan (WithholdingDate)

    for r in range(data_start_row, ws.max_row + 1):
        # Check if row has data
        masa = ws.cell(row=r, column=2).value
        npwp = ws.cell(row=r, column=5).value
        if masa is None and npwp is None:
            continue

        tax_object_code = _safe_str(ws.cell(row=r, column=10).value)
        ptkp_status = _safe_str(ws.cell(row=r, column=7).value)
        gross = _safe_float(ws.cell(row=r, column=11).value)

        # Calculate rate — always recalculate
        rate_cell = ws.cell(row=r, column=12).value
        if _is_formula(rate_cell) or rate_cell is None:
            rate = calculate_rate(tax_object_code, ptkp_status, gross, 100, "BPMP")
        else:
            rate = _safe_float(rate_cell)

        row_data = {
            "TaxPeriodMonth": int(_safe_float(masa)) if masa else 0,
            "TaxPeriodYear": int(_safe_float(ws.cell(row=r, column=3).value)) if ws.cell(row=r, column=3).value else 0,
            "CounterpartOpt": _safe_str(ws.cell(row=r, column=4).value) or "Resident",
            "CounterpartTin": _safe_str(npwp),
            "CounterpartPassport": _safe_str(ws.cell(row=r, column=6).value) or None,
            "StatusTaxExemption": ptkp_status,
            "Position": _safe_str(ws.cell(row=r, column=8).value),
            "TaxCertificate": _safe_str(ws.cell(row=r, column=9).value),
            "TaxObjectCode": tax_object_code,
            "Gross": gross,
            "Rate": rate,
            "IDPlaceOfBusinessActivity": _safe_str(ws.cell(row=r, column=13).value),
            "WithholdingDate": _safe_date(ws.cell(row=r, column=14).value),
        }

        # Validate
        row_num = r
        if len(row_data["CounterpartTin"]) != 16 and row_data["CounterpartTin"]:
            errors.append(f"Baris {row_num}: NPWP/NIK harus 16 digit (ditemukan: {len(row_data['CounterpartTin'])} digit)")
        if tax_object_code and tax_object_code not in BPMP_TAX_OBJECT_LOOKUP:
            errors.append(f"Baris {row_num}: Kode objek pajak '{tax_object_code}' tidak dikenali untuk BPMP")

        rows.append(row_data)

    if not rows:
        errors.append("Tidak ada data yang ditemukan di sheet DATA.")

    return {"tin": tin, "rows": rows, "errors": errors}


def read_bpa1(file_bytes: bytes) -> dict:
    """
    Read BPA1 Excel template (Bukti Potong Tahunan A1).

    Expected structure:
    - Sheet "DATA" with Table1 (range B3:AB)
    - Row 1: A1:B1 merged = label "NPWP Pemotong", C1 = NPWP value
    - Row 3: column headers
    - Row 4+: data rows (columns B-AB = 27 fields)

    Returns:
        dict with tin, rows, errors
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=False)

    if "DATA" not in wb.sheetnames:
        return {"tin": "", "rows": [], "errors": ["Sheet 'DATA' tidak ditemukan dalam file Excel."]}

    ws = wb["DATA"]
    errors = []
    rows = []

    # TIN from cell C1
    tin = _safe_str(ws.cell(row=1, column=3).value)
    if not tin:
        for r in range(1, 4):
            for c in range(1, 6):
                val = _safe_str(ws.cell(row=r, column=c).value)
                if len(val) == 16 and val.isdigit():
                    tin = val
                    break
            if tin:
                break

    if not tin:
        errors.append("NPWP Pemotong tidak ditemukan (diharapkan di cell C1).")

    # Column mapping for BPA1 DATA sheet:
    # B(2)=Pemberi Kerja Selanjutnya, C(3)=Masa Pajak Awal, D(4)=Masa Pajak Akhir,
    # E(5)=Tahun Pajak, F(6)=WNI/WNA, G(7)=No. Paspor, H(8)=NPWP,
    # I(9)=Status PTKP, J(10)=Posisi, K(11)=Kode Objek Pajak,
    # L(12)=Status Bukti Potong, M(13)=Jumlah Bulan Bekerja,
    # N(14)=Gaji, O(15)=Opsi Gross Up, P(16)=Tunjangan PPh,
    # Q(17)=Tunjangan Lainnya, R(18)=Honorarium, S(19)=Asuransi,
    # T(20)=Natura, U(21)=Tantiem/Bonus/THR, V(22)=Iuran Pensiun/THT/JHT,
    # W(23)=Zakat, X(24)=Nomor BP Sebelumnya, Y(25)=Fasilitas Pajak,
    # Z(26)=PPh Pasal 21, AA(27)=ID TKU Pemotong, AB(28)=Tanggal Pemotongan

    data_start_row = 4
    for r in range(data_start_row, ws.max_row + 1):
        npwp = ws.cell(row=r, column=8).value
        masa_start = ws.cell(row=r, column=3).value
        if npwp is None and masa_start is None:
            continue

        passport_val = _safe_str(ws.cell(row=r, column=7).value)
        prev_wh_val = _safe_str(ws.cell(row=r, column=24).value)

        row_data = {
            "WorkForSecondEmployer": _safe_str(ws.cell(row=r, column=2).value) or "No",
            "TaxPeriodMonthStart": int(_safe_float(masa_start)) if masa_start else 0,
            "TaxPeriodMonthEnd": int(_safe_float(ws.cell(row=r, column=4).value)) if ws.cell(row=r, column=4).value else 0,
            "TaxPeriodYear": int(_safe_float(ws.cell(row=r, column=5).value)) if ws.cell(row=r, column=5).value else 0,
            "CounterpartOpt": _safe_str(ws.cell(row=r, column=6).value) or "Resident",
            "CounterpartPassport": passport_val if passport_val else None,
            "CounterpartTin": _safe_str(npwp),
            "TaxExemptOpt": _safe_str(ws.cell(row=r, column=9).value),
            "CounterpartPosition": _safe_str(ws.cell(row=r, column=10).value),
            "TaxObjectCode": _safe_str(ws.cell(row=r, column=11).value),
            "StatusOfWithholding": _safe_str(ws.cell(row=r, column=12).value),
            "NumberOfMonths": int(_safe_float(ws.cell(row=r, column=13).value)),
            "SalaryPensionJhtTht": _safe_float(ws.cell(row=r, column=14).value),
            "GrossUpOpt": _safe_str(ws.cell(row=r, column=15).value) or "No",
            "IncomeTaxBenefit": _safe_float(ws.cell(row=r, column=16).value),
            "OtherBenefit": _safe_float(ws.cell(row=r, column=17).value),
            "Honorarium": _safe_float(ws.cell(row=r, column=18).value),
            "InsurancePaidByEmployer": _safe_float(ws.cell(row=r, column=19).value),
            "Natura": _safe_float(ws.cell(row=r, column=20).value),
            "TantiemBonusThr": _safe_float(ws.cell(row=r, column=21).value),
            "PensionContributionJhtThtFee": _safe_float(ws.cell(row=r, column=22).value),
            "Zakat": _safe_float(ws.cell(row=r, column=23).value),
            "PrevWhTaxSlip": prev_wh_val if prev_wh_val else None,
            "TaxCertificate": _safe_str(ws.cell(row=r, column=25).value) or "N/A",
            "Article21IncomeTax": _safe_float(ws.cell(row=r, column=26).value),
            "IDPlaceOfBusinessActivity": _safe_str(ws.cell(row=r, column=27).value),
            "WithholdingDate": _safe_date(ws.cell(row=r, column=28).value),
        }

        # Validate
        if len(row_data["CounterpartTin"]) != 16 and row_data["CounterpartTin"]:
            errors.append(f"Baris {r}: NPWP harus 16 digit (ditemukan: {len(row_data['CounterpartTin'])} digit)")
        if row_data["TaxObjectCode"] and row_data["TaxObjectCode"] not in BPA1_TAX_OBJECT_LOOKUP:
            errors.append(f"Baris {r}: Kode objek pajak '{row_data['TaxObjectCode']}' tidak dikenali untuk BPA1")

        rows.append(row_data)

    if not rows:
        errors.append("Tidak ada data yang ditemukan di sheet DATA.")

    return {"tin": tin, "rows": rows, "errors": errors}


def read_bppu(file_bytes: bytes) -> dict:
    """
    Read BPPU Excel template (Bukti Potong PPh Unifikasi).

    Expected structure:
    - Sheet "DATA" with Table1 (range B3:P)
    - Row 1: A1:B1 merged = label "NPWP Pemotong", C1 = NPWP value
    - Row 3: column headers
    - Row 4+: data rows (columns B-P = 15 fields)

    Returns:
        dict with tin, rows, errors
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=False)

    if "DATA" not in wb.sheetnames:
        return {"tin": "", "rows": [], "errors": ["Sheet 'DATA' tidak ditemukan dalam file Excel."]}

    ws = wb["DATA"]
    errors = []
    rows = []

    # TIN from cell C1
    tin = _safe_str(ws.cell(row=1, column=3).value)
    if not tin:
        for r in range(1, 4):
            for c in range(1, 6):
                val = _safe_str(ws.cell(row=r, column=c).value)
                if len(val) == 16 and val.isdigit():
                    tin = val
                    break
            if tin:
                break

    if not tin:
        errors.append("NPWP Pemotong tidak ditemukan (diharapkan di cell C1).")

    # Column mapping for BPPU DATA sheet:
    # B(2)=Masa Pajak, C(3)=Tahun Pajak, D(4)=NPWP, E(5)=ID TKU Penerima,
    # F(6)=Fasilitas, G(7)=Kode Objek Pajak, H(8)=DPP, I(9)=Tarif (formula),
    # J(10)=Jenis Dok Referensi, K(11)=Nomor Dok, L(12)=Tanggal Dok,
    # M(13)=ID TKU Pemotong, N(14)=Opsi Pembayaran (IP), O(15)=Nomor SP2D,
    # P(16)=Tanggal Pemotongan

    data_start_row = 4
    for r in range(data_start_row, ws.max_row + 1):
        masa = ws.cell(row=r, column=2).value
        npwp = ws.cell(row=r, column=4).value
        if masa is None and npwp is None:
            continue

        tax_object_code = _safe_str(ws.cell(row=r, column=7).value)

        # Calculate rate — VLOOKUP from reference table
        rate_cell = ws.cell(row=r, column=9).value
        if _is_formula(rate_cell) or rate_cell is None:
            rate = get_bppu_rate(tax_object_code)
        else:
            rate = _safe_float(rate_cell)

        sp2d_val = _safe_str(ws.cell(row=r, column=15).value)

        row_data = {
            "TaxPeriodMonth": int(_safe_float(masa)) if masa else 0,
            "TaxPeriodYear": int(_safe_float(ws.cell(row=r, column=3).value)) if ws.cell(row=r, column=3).value else 0,
            "CounterpartTin": _safe_str(npwp),
            "IDPlaceOfBusinessActivityOfIncomeRecipient": _safe_str(ws.cell(row=r, column=5).value),
            "TaxCertificate": _safe_str(ws.cell(row=r, column=6).value) or "N/A",
            "TaxObjectCode": tax_object_code,
            "TaxBase": _safe_float(ws.cell(row=r, column=8).value),
            "Rate": rate,
            "Document": _safe_str(ws.cell(row=r, column=10).value),
            "DocumentNumber": _safe_str(ws.cell(row=r, column=11).value),
            "DocumentDate": _safe_date(ws.cell(row=r, column=12).value),
            "IDPlaceOfBusinessActivity": _safe_str(ws.cell(row=r, column=13).value),
            "GovTreasurerOpt": _safe_str(ws.cell(row=r, column=14).value) or "N/A",
            "SP2DNumber": sp2d_val if sp2d_val else None,
            "WithholdingDate": _safe_date(ws.cell(row=r, column=16).value),
        }

        # Validate
        if len(row_data["CounterpartTin"]) != 16 and row_data["CounterpartTin"]:
            errors.append(f"Baris {r}: NPWP harus 16 digit (ditemukan: {len(row_data['CounterpartTin'])} digit)")
        if tax_object_code and tax_object_code not in BPPU_TAX_OBJECT_LOOKUP:
            errors.append(f"Baris {r}: Kode objek pajak '{tax_object_code}' tidak dikenali untuk BPPU")

        rows.append(row_data)

    if not rows:
        errors.append("Tidak ada data yang ditemukan di sheet DATA.")

    return {"tin": tin, "rows": rows, "errors": errors}
